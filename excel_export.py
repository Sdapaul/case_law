"""
판례 목록 → Excel(.xlsx) 변환
"""

import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)

logger = logging.getLogger(__name__)

_HEADERS = [
    "발송일", "판례명", "사건번호", "법원", "사건유형",
    "선고일", "AI요약", "참조 URL", "우선순위여부", "최신판례여부",
]

_COL_WIDTHS = [12, 45, 22, 18, 14, 12, 60, 55, 12, 12]

_HEAD_FILL  = PatternFill("solid", fgColor="1A56C4")
_PRIO_FILL  = PatternFill("solid", fgColor="DDEAFF")
_FRESH_FILL = PatternFill("solid", fgColor="E6FFE6")
_ALT_FILL   = PatternFill("solid", fgColor="F2F6FF")

_HEAD_FONT  = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=11)
_BODY_FONT  = Font(name="맑은 고딕", size=10)
_LINK_FONT  = Font(name="맑은 고딕", size=10, color="1A56C4", underline="single")

_THIN = Side(style="thin", color="D0D8EE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP   = Alignment(vertical="top", wrap_text=True)


def _add_header(ws) -> None:
    ws.append(_HEADERS)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill   = _HEAD_FILL
        cell.font   = _HEAD_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER
        ws.column_dimensions[cell.column_letter].width = _COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def generate_excel(cases: list[dict], title: str = "판례 목록") -> bytes:
    """판례 목록을 Excel 바이트로 변환"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    _add_header(ws)

    for row_idx, c in enumerate(cases, 2):
        is_priority = bool(c.get("is_priority"))
        is_fresh    = bool(c.get("is_fresh"))
        sent_date   = c.get("sent_date", "")
        link        = c.get("link", "")

        row_data = [
            sent_date,
            c.get("title", ""),
            c.get("case_num", ""),
            c.get("court", ""),
            c.get("case_type", ""),
            c.get("decision_date", "") or c.get("date", ""),
            c.get("ai_summary", "") or c.get("summary", ""),
            link,
            "★ 우선순위" if is_priority else "",
            "✓ 최신" if is_fresh else "",
        ]
        ws.append(row_data)

        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border    = _BORDER
            cell.alignment = _CENTER if col_idx in (1, 3, 4, 5, 6, 9, 10) else _WRAP

            if col_idx == 8 and link:
                cell.font = _LINK_FONT
                cell.hyperlink = link
            else:
                cell.font = _BODY_FONT

        # 행 배경색: 우선순위 > 최신 > 짝수행 교차
        fill = (
            _PRIO_FILL  if is_priority else
            _FRESH_FILL if is_fresh    else
            _ALT_FILL   if row_idx % 2 == 0 else None
        )
        if fill:
            for cell in ws[row_idx]:
                if cell.column != 8:
                    cell.fill = fill

        ws.row_dimensions[row_idx].height = 60

    # 요약 열 너비 최종 조정
    ws.column_dimensions["G"].width = 60

    output = io.BytesIO()
    wb.save(output)
    logger.info(f"Excel 생성 완료: {len(cases)}건")
    return output.getvalue()
